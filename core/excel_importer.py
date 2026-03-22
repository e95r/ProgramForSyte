from __future__ import annotations

import re
from zipfile import BadZipFile
from pathlib import Path

from core.time_utils import parse_seed_time_to_cs

HEADER_ALIASES = {
    "name": {"фи", "ф. и.", "фио", "участник", "name"},
    "year": {"год рождения", "год", "birth", "year"},
    "team": {"команда", "team"},
    "seed": {"заявочное время", "время", "seed", "entry time"},
    "heat_lane": {"заплыв/дорожка", "заплыв/дорожка ", "заплыв", "heat/lane"},
    "heat": {"заплыв", "heat"},
    "lane": {"дорожка", "lane"},
}

EVENT_TITLE_RE = re.compile(
    r"^\s*(?:\d+\s*[xх×]\s*\d+|\d+)(?:\s*[-–—]\s*|\s*(?:м|метр|метров)\b\s*).+$",
    re.IGNORECASE,
)


class ExcelImportError(ValueError):
    """Raised when a startlist file cannot be parsed as supported Excel."""


def _normalize(value: object) -> str:
    return str(value or "").strip().lower()


def _find_columns(header_row: list[object]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for idx, raw in enumerate(header_row):
        h = _normalize(raw)
        for key, aliases in HEADER_ALIASES.items():
            if h in aliases and key not in mapping:
                mapping[key] = idx
    return mapping


def _parse_heat_lane(text: object) -> tuple[int | None, int | None]:
    raw = str(text or "").strip().replace(" ", "")
    if not raw or "/" not in raw:
        return None, None
    a, b = raw.split("/", 1)
    if a.isdigit() and b.isdigit():
        return int(a), int(b)
    return None, None


def _parse_birth_year(value: object) -> int | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        year = int(value)
        return year if 1900 <= year <= 2100 else None
    text = str(value).strip()
    if not text:
        return None
    digits = "".join(ch for ch in text if ch.isdigit())
    if len(digits) == 4:
        year = int(digits)
        return year if 1900 <= year <= 2100 else None
    return None


def _row_non_empty_values(row: tuple[object, ...]) -> list[object]:
    return [value for value in row if str(value or "").strip()]


def _looks_like_header_row(row: tuple[object, ...]) -> bool:
    cols = _find_columns(list(row))
    return "name" in cols and (
        "heat_lane" in cols or ("heat" in cols and "lane" in cols) or "year" in cols or "seed" in cols
    )


def _looks_like_event_title(row: tuple[object, ...]) -> bool:
    values = _row_non_empty_values(row)
    if len(values) != 1:
        return False
    text = str(values[0]).strip()
    if not text or _looks_like_header_row(row):
        return False
    return bool(EVENT_TITLE_RE.match(text))


def _extract_event_title(row: tuple[object, ...], fallback: str) -> str:
    values = _row_non_empty_values(row)
    if len(values) == 1:
        return str(values[0]).strip()
    return fallback


def _parse_separate_heat_lane(
    row: tuple[object, ...],
    cols: dict[str, int],
    current_heat: int | None,
) -> tuple[int | None, int | None]:
    heat = current_heat
    if "heat" in cols:
        raw_heat = row[cols["heat"]]
        if isinstance(raw_heat, (int, float)) and int(raw_heat) > 0:
            heat = int(raw_heat)
        else:
            heat_text = str(raw_heat or "").strip()
            if heat_text.isdigit():
                heat = int(heat_text)
    lane = None
    if "lane" in cols:
        raw_lane = row[cols["lane"]]
        if isinstance(raw_lane, (int, float)) and int(raw_lane) > 0:
            lane = int(raw_lane)
        else:
            lane_text = str(raw_lane or "").strip()
            if lane_text.isdigit():
                lane = int(lane_text)
    return heat, lane


def _parse_swimmers(rows: list[tuple[object, ...]], start_idx: int, cols: dict[str, int]) -> list[dict]:
    swimmers: list[dict] = []
    current_heat: int | None = None
    for row in rows[start_idx:]:
        if _looks_like_header_row(row):
            break
        if _looks_like_event_title(row):
            break
        if not _row_non_empty_values(row):
            continue

        name = str(row[cols["name"]] or "").strip()
        if not name:
            if "heat" in cols and "name" in cols:
                heat_candidate, _ = _parse_separate_heat_lane(row, cols, current_heat)
                current_heat = heat_candidate
            continue

        year = row[cols["year"]] if "year" in cols else None
        team = row[cols["team"]] if "team" in cols else None
        seed_raw = row[cols["seed"]] if "seed" in cols else None
        seed_raw_text = str(seed_raw).strip() if seed_raw is not None else None

        heat, lane = (None, None)
        if "heat_lane" in cols and "lane" not in cols:
            heat, lane = _parse_heat_lane(row[cols["heat_lane"]])
        if (heat is None or lane is None) and "lane" in cols:
            heat, lane = _parse_separate_heat_lane(row, cols, current_heat)
        current_heat = heat or current_heat
        source_heat_lane = "separate" if "lane" in cols else "combined" if "heat_lane" in cols else None

        swimmers.append(
            {
                "full_name": name,
                "birth_year": _parse_birth_year(year),
                "team": str(team).strip() if team is not None else None,
                "seed_time_raw": seed_raw_text,
                "seed_time_cs": parse_seed_time_to_cs(seed_raw_text),
                "heat": heat,
                "lane": lane,
                "source_heat_lane": source_heat_lane,
                "status": "OK",
            }
        )
    return swimmers


def import_excel(path: Path) -> dict[str, list[dict]]:
    def _file_debug_message(file_path: Path) -> str:
        exists = file_path.exists()
        size = file_path.stat().st_size if exists else 0
        suffix = file_path.suffix.lower()
        return (
            f"Selected: {file_path}\n"
            f"Exists: {exists}\n"
            f"Size: {size}\n"
            f"Suffix: {suffix}"
        )

    def _validate_input_file(file_path: Path) -> None:
        if not file_path.exists():
            raise ExcelImportError("Выбранный файл не существует.")
        size = file_path.stat().st_size
        suffix = file_path.suffix.lower()
        if size == 0:
            raise ExcelImportError("Выбранный файл пустой (0 байт). Выберите корректный Excel-файл.")
        if suffix not in {".xlsx", ".xlsm"}:
            raise ExcelImportError("Поддерживаются только файлы .xlsx и .xlsm.")

    initial_debug_message = _file_debug_message(path)
    print(initial_debug_message)
    _validate_input_file(path)

    if path.suffix.lower() == ".xls":
        raise ExcelImportError("Формат .xls не поддерживается. Сохраните файл как .xlsx и попробуйте снова.")

    try:
        from openpyxl import load_workbook
        from openpyxl.utils.exceptions import InvalidFileException
    except ModuleNotFoundError as exc:
        raise ExcelImportError("Не установлен пакет openpyxl. Установите зависимости приложения.") from exc

    try:
        pre_load_debug_message = _file_debug_message(path)
        print(pre_load_debug_message)
        _validate_input_file(path)
        wb = load_workbook(path, data_only=True)
    except (BadZipFile, InvalidFileException) as exc:
        raise ExcelImportError(
            "Не удалось открыть Excel-файл. Проверьте, что это корректный .xlsx/.xlsm файл."
        ) from exc

    result: dict[str, list[dict]] = {}

    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        pending_event_title = ws.title
        for idx, row in enumerate(rows):
            if _looks_like_event_title(row):
                pending_event_title = _extract_event_title(row, ws.title)
                continue

            cols = _find_columns(list(row))
            if "name" not in cols:
                continue

            swimmers = _parse_swimmers(rows, idx + 1, cols)
            if swimmers:
                event_title = pending_event_title or ws.title
                suffix = 2
                original_title = event_title
                while event_title in result:
                    event_title = f"{original_title} ({suffix})"
                    suffix += 1
                result[event_title] = swimmers
                pending_event_title = ws.title

    return result
