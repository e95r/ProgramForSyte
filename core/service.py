from __future__ import annotations

import hashlib
import html
import json
import re
import secrets
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from core.db import MeetRepository
from core.models import Secretary
from core.reseeding import compress_lanes_within_heats, full_reseed
from core.time_utils import parse_seed_time_to_cs


EVENT_NAME_PARTS_RE = re.compile(
    r"^\s*(?P<base>.+?)(?:\s*,\s*(?P<gender>женщины|девушки|девочки|мужчины|юноши|мальчики|все))?(?:\s+(?P<age>все))?\s*$",
    re.IGNORECASE,
)
TRAILING_ALL_RE = re.compile(r"^(?P<title>.+?)(?:\s*,)?\s+все\s*$", re.IGNORECASE)


class MeetService:
    def __init__(self, root: Path):
        self.root = root
        self.meet_dir = self.root / "meet"
        self.backup_dir = self.meet_dir / "backups"
        self.db_path = self.meet_dir / "meet.db"
        self.backup_dir.mkdir(parents=True, exist_ok=True)
        if self.db_path.exists():
            self.create_backup(reason="startup")
        self.repo = MeetRepository(self.db_path)

    def close(self) -> None:
        self.repo.close()

    def create_backup(self, reason: str = "manual") -> Path | None:
        if not self.db_path.exists():
            return None
        stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
        backup_path = self.backup_dir / f"meet-{reason}-{stamp}.db"
        shutil.copy2(self.db_path, backup_path)
        return backup_path

    def secretary_count(self) -> int:
        return self.repo.secretary_count()

    def _hash_password(self, password: str, salt: str | None = None) -> str:
        salt = salt or secrets.token_hex(16)
        digest = hashlib.sha256(f"{salt}:{password}".encode("utf-8")).hexdigest()
        return f"{salt}${digest}"

    def _verify_password(self, password: str, stored_hash: str) -> bool:
        try:
            salt, digest = stored_hash.split("$", 1)
        except ValueError:
            return False
        return self._hash_password(password, salt) == f"{salt}${digest}"

    def register_secretary(
        self,
        username: str,
        password: str,
        password_hint: str,
        display_name: str = "",
    ) -> Secretary:
        username = username.strip()
        display_name = display_name.strip() or username
        password_hint = password_hint.strip()
        if not username:
            raise ValueError("Укажите логин секретаря")
        if len(password) < 4:
            raise ValueError("Пароль должен содержать минимум 4 символа")
        if not password_hint:
            raise ValueError("Укажите подсказку для восстановления пароля")
        if self.repo.get_secretary_by_username(username) is not None:
            raise ValueError("Секретарь с таким логином уже зарегистрирован")
        secretary_id = self.repo.create_secretary(
            username=username,
            display_name=display_name,
            password_hash=self._hash_password(password),
            password_hint=password_hint,
        )
        self.repo.log("register_secretary", f"username={username}")
        return Secretary(secretary_id, username, display_name, password_hint)

    def authenticate_secretary(self, username: str, password: str) -> Secretary | None:
        row = self.repo.get_secretary_auth_row(username.strip())
        if row is None or not self._verify_password(password, row["password_hash"]):
            return None
        secretary = Secretary(
            id=int(row["id"]),
            username=row["username"],
            display_name=row["display_name"],
            password_hint=row["password_hint"],
        )
        self.repo.log("authenticate_secretary", f"username={secretary.username}")
        return secretary

    def get_secretary_password_hint(self, username: str) -> str | None:
        secretary = self.repo.get_secretary_by_username(username.strip())
        if secretary is None:
            return None
        self.repo.log("password_hint_requested", f"username={secretary.username}")
        return secretary.password_hint

    def import_startlist(self, excel_path: Path) -> None:
        self.repo.clear_all()
        from core.excel_importer import extract_meet_metadata, import_excel

        imported = import_excel(excel_path)
        metadata = extract_meet_metadata(excel_path)
        self.repo.set_meta("competition_title", metadata.get("competition_title") or self._derive_competition_title(excel_path))
        self.repo.set_meta("competition_date", metadata.get("competition_date") or "")
        self.repo.set_meta("competition_place", metadata.get("competition_place") or "")
        self.repo.set_meta("age_groups", json.dumps(metadata.get("age_groups") or [], ensure_ascii=False))
        self.repo.set_meta("relay_age_groups", json.dumps(metadata.get("relay_age_groups") or [], ensure_ascii=False))
        for event_name, swimmers in imported.items():
            normalized_event_name = self._normalize_imported_event_name(event_name)
            lanes_count = self._infer_imported_lanes_count(swimmers)
            event_id = self.repo.upsert_event(normalized_event_name, lanes_count=lanes_count)
            swimmers = self._normalize_imported_start_protocol(swimmers, lanes_count=lanes_count)
            self.repo.add_swimmers(event_id, swimmers)
        self.repo.log("import_excel", str(excel_path))

    def _derive_competition_title(self, excel_path: Path) -> str:
        raw_title = excel_path.stem.replace("_", " ").replace("-", " ")
        normalized = " ".join(raw_title.split())
        return normalized or "Итоговый протокол соревнований"

    def _normalize_imported_event_name(self, event_name: str) -> str:
        normalized = " ".join(event_name.split()).strip(" ,")
        match = TRAILING_ALL_RE.match(normalized)
        if match:
            normalized = match.group("title").strip(" ,")
        return normalized or event_name.strip()


    def _infer_imported_lanes_count(self, swimmers: list[dict], default: int = 8) -> int:
        lanes = [int(lane) for lane in (s.get("lane") for s in swimmers) if isinstance(lane, int) and lane > 0]
        if lanes:
            return max(lanes)

        heats: dict[int, int] = {}
        for swimmer in swimmers:
            heat = swimmer.get("heat")
            if isinstance(heat, int) and heat > 0:
                heats[heat] = heats.get(heat, 0) + 1
        if heats:
            return max(heats.values())

        return default

    def _normalize_imported_start_protocol(self, swimmers: list[dict], lanes_count: int = 8) -> list[dict]:
        active = [dict(s) for s in swimmers if s.get("status") != "DNS"]
        if active and all(s.get("heat") is not None and s.get("lane") is not None for s in active):
            ordered_active = sorted(
                active,
                key=lambda s: (
                    s.get("heat") is None,
                    s.get("heat") or 10**12,
                    s.get("lane") is None,
                    s.get("lane") or 10**12,
                    (s.get("full_name") or "").lower(),
                ),
            )
            dns = [dict(s) for s in swimmers if s.get("status") == "DNS"]
            for swimmer in dns:
                swimmer["heat"] = None
                swimmer["lane"] = None
            return ordered_active + dns
        return self._rebuild_start_protocol(swimmers, lanes_count=lanes_count)

    def _rebuild_start_protocol(self, swimmers: list[dict], lanes_count: int = 8) -> list[dict]:
        active = [dict(s) for s in swimmers if s.get("status") != "DNS"]
        dns = [dict(s) for s in swimmers if s.get("status") == "DNS"]
        active.sort(
            key=lambda s: (
                s.get("seed_time_cs") is None,
                s.get("seed_time_cs") or 10**12,
                (s.get("full_name") or "").lower(),
            )
        )
        for idx, swimmer in enumerate(active):
            swimmer["heat"] = idx // lanes_count + 1
            swimmer["lane"] = idx % lanes_count + 1
        for swimmer in dns:
            swimmer["heat"] = None
            swimmer["lane"] = None
        return active + dns

    def mark_dns(self, event_id: int, swimmer_ids: list[int]) -> None:
        self.repo.set_dns(swimmer_ids)
        self.repo.log("mark_dns", f"event={event_id}; ids={swimmer_ids}")

    def restore_swimmers(self, event_id: int, swimmer_ids: list[int], mode: str = "soft") -> None:
        self.repo.restore_swimmers(swimmer_ids)
        self.reseed_event(event_id, mode=mode)
        self.repo.log("restore_swimmers", f"event={event_id}; ids={swimmer_ids}; mode={mode}")

    def reseed_event(self, event_id: int, mode: str = "soft") -> None:
        event = next(e for e in self.repo.list_events() if e.id == event_id)
        swimmers = self.repo.list_swimmers(event_id)
        if mode == "full":
            updated = full_reseed(swimmers, lanes_count=event.lanes_count)
        else:
            updated = compress_lanes_within_heats(swimmers, lanes_count=event.lanes_count)
        self.repo.update_swimmer_positions(updated)
        self.repo.log("reseed_event", f"event={event_id}; mode={mode}")

    def save_event_results(self, event_id: int, results: list[dict[str, str]]) -> None:
        payload: list[tuple[int, str | None, int | None, str | None]] = []
        for row in results:
            swimmer_id = int(row["swimmer_id"])
            raw = row.get("result_time_raw", "").strip() or None
            mark = row.get("result_mark", "").strip() or None
            cs = parse_seed_time_to_cs(raw) if raw else None
            payload.append((swimmer_id, raw, cs, mark))
        self.repo.save_results(payload)
        self.repo.log("save_event_results", f"event={event_id}; rows={len(payload)}")

    def build_event_protocol(
        self,
        event_id: int,
        grouped: bool = True,
        sort_by: str = "place",
        sort_desc: bool = False,
        group_by: str = "heat",
    ) -> str:
        event = next(e for e in self.repo.list_events() if e.id == event_id)
        swimmers = self.repo.list_swimmers(event_id)
        results_mode = any(self._swimmer_has_result(swimmer) for swimmer in swimmers)
        return self._build_protocol_document(
            page_title=self.repo.get_meta("competition_title") or event.name,
            events=[(event.name, swimmers)],
            final_mode=results_mode,
            sort_by=sort_by,
            sort_desc=sort_desc,
            group_by=group_by,
            grouped=grouped,
            show_heat_column=False,
        )

    def build_final_protocol(
        self,
        grouped: bool = True,
        sort_by: str = "place",
        sort_desc: bool = False,
        group_by: str = "heat",
    ) -> str:
        events = []
        for event in self.repo.list_events():
            swimmers = self._filter_final_protocol_swimmers(self.repo.list_swimmers(event.id))
            events.append((event.name, swimmers))
        return self._build_protocol_document(
            page_title=f"Итоговый протокол {self.repo.get_meta('competition_title') or 'соревнований'}",
            events=events,
            final_mode=True,
            sort_by=sort_by,
            sort_desc=sort_desc,
            group_by=group_by,
            grouped=grouped,
            show_heat_column=False,
        )

    def _build_protocol_document(
        self,
        page_title: str,
        events: list[tuple[str, list]],
        final_mode: bool,
        sort_by: str,
        sort_desc: bool,
        group_by: str,
        grouped: bool,
        show_heat_column: bool,
    ) -> str:
        title = self.repo.get_meta("competition_title") or "Итоговый протокол соревнований"
        date = self.repo.get_meta("competition_date") or ""
        place = self.repo.get_meta("competition_place") or ""
        body: list[str] = [self._protocol_styles(), "<div class='page'>"]
        doc_title = title if not final_mode else page_title
        body.append(f"<div class='doc-title'>{html.escape(doc_title)}</div>")
        body.append("<table class='meta-table'>")
        body.append(f"<tr><td>{html.escape(date)}</td></tr>")
        body.append(f"<tr><td>{html.escape(place)}</td></tr>")
        body.append("</table>")

        for event_name, swimmers in events:
            groups = self._split_event_into_age_groups(event_name, swimmers)
            for group in groups:
                body.append(
                    self._build_age_group_table_html(
                        event_name=event_name,
                        swimmers=group["swimmers"],
                        age_label=group["age_label"],
                        gender_label=group["gender_label"],
                        gender_color=group["gender_color"],
                        final_mode=final_mode,
                        sort_by=sort_by,
                        sort_desc=sort_desc,
                        group_by=group_by,
                        grouped=grouped,
                        show_heat_column=show_heat_column,
                    )
                )
        body.append("</div>")
        return "\n".join(body)

    def _protocol_styles(self) -> str:
        return (
            "<style>"
            "@page { size: A4 portrait; margin: 12mm; }"
            "html, body { margin: 0; padding: 0; }"
            "body { font-family: Arial, sans-serif; color: #000; }"
            ".page { width: 186mm; max-width: 186mm; margin: 0 auto; box-sizing: border-box; }"
            ".doc-title { width: 100%; text-align: center; font-size: 24px; font-weight: 700; margin: 0 0 18px 0; }"
            ".meet-line { font-size: 16px; margin: 0 0 8px 0; }"
            ".meta-table { width: 100%; border-collapse: collapse; margin: 0 0 16px 0; table-layout: fixed; }"
            ".meta-table td { border: 1px solid #cfcfcf; padding: 8px 10px; text-align: center; font-size: 16px; }"
            ".protocol-table { width: 100%; border-collapse: collapse; margin: 0 0 18px 0; font-size: 14px; table-layout: fixed; }"
            ".protocol-table th, .protocol-table td { border: 1px solid #cfcfcf; padding: 6px 8px; }"
            ".protocol-table th { text-align: left; }"
            ".protocol-table .col-1 { width: 11%; }"
            ".protocol-table .col-2 { width: 11%; }"
            ".protocol-table .col-3 { width: 28%; }"
            ".protocol-table .col-4 { width: 14%; }"
            ".protocol-table .col-5 { width: 24%; }"
            ".protocol-table .col-6 { width: 12%; }"
            ".category-title { color: #fff; text-align: center; font-weight: 700; font-size: 18px; text-transform: uppercase; padding: 12px 8px; }"
            ".category-title.boys { background: #5b9bd5; }"
            ".category-title.girls { background: #e06666; }"
            ".category-title.mixed { background: #808080; }"
            ".num, .place, .year, .heat, .lane, .time { text-align: center; }"
            ".name { text-align: left; }"
            ".heat-label td { background: #f1f1f1; font-weight: 700; text-align: center; }"
            ".heat-merged { vertical-align: middle; font-weight: 700; }"
            "@media print { .page { width: 100%; max-width: none; margin: 0; } }"
            "</style>"
        )

    def _split_event_into_age_groups(self, event_name: str, swimmers: list) -> list[dict]:
        age_groups = self._load_age_groups(event_name)
        gender_label, gender_color = self._detect_gender(event_name)
        base_title = self._base_event_name(event_name)
        if not age_groups:
            return [{
                "base_title": base_title,
                "age_label": "Все возраста",
                "gender_label": gender_label,
                "gender_color": gender_color,
                "swimmers": list(swimmers),
            }]

        result: list[dict] = []
        for group in age_groups:
            filtered = [s for s in swimmers if self._matches_age_group(s.birth_year, group)]
            if filtered:
                result.append(
                    {
                        "base_title": base_title,
                        "age_label": group["label"],
                        "gender_label": gender_label,
                        "gender_color": gender_color,
                        "swimmers": filtered,
                    }
                )
        if result:
            return result
        return [{
            "base_title": base_title,
            "age_label": "Все возраста",
            "gender_label": gender_label,
            "gender_color": gender_color,
            "swimmers": list(swimmers),
        }]

    def _load_age_groups(self, event_name: str) -> list[dict]:
        raw = self.repo.get_meta("relay_age_groups" if self._is_relay_event(event_name) else "age_groups") or "[]"
        try:
            data = json.loads(raw)
        except json.JSONDecodeError:
            return []
        return [group for group in data if isinstance(group, dict) and group.get("label")]

    def _is_relay_event(self, event_name: str) -> bool:
        lowered = event_name.lower()
        return "эстаф" in lowered or "x" in lowered or "х" in lowered or "×" in lowered

    def _matches_age_group(self, birth_year: int | None, group: dict) -> bool:
        if birth_year is None:
            return False
        min_year = group.get("min_year")
        max_year = group.get("max_year")
        if min_year is not None and birth_year < int(min_year):
            return False
        if max_year is not None and birth_year > int(max_year):
            return False
        return True

    def _parse_event_name(self, event_name: str) -> tuple[str, str]:
        normalized = " ".join(event_name.split())
        match = EVENT_NAME_PARTS_RE.match(normalized)
        if not match:
            return event_name.strip(), "Все"

        base_name = (match.group("base") or "").strip(" ,") or event_name.strip()
        gender_token = (match.group("gender") or "").lower()
        age_token = (match.group("age") or "").lower()

        if gender_token in {"женщины", "девушки", "девочки"}:
            return base_name, "Женщины"
        if gender_token in {"мужчины", "юноши", "мальчики"}:
            return base_name, "Мужчины"
        if gender_token == "все" or age_token == "все":
            return base_name, "Все"
        return base_name, "Все"

    def _detect_gender(self, event_name: str) -> tuple[str, str]:
        _, gender_label = self._parse_event_name(event_name)
        if gender_label == "Женщины":
            return gender_label, "girls"
        if gender_label == "Мужчины":
            return gender_label, "boys"
        return gender_label, "mixed"

    def _base_event_name(self, event_name: str) -> str:
        base_name, _ = self._parse_event_name(event_name)
        return base_name

    def export_event_protocol_excel(
        self,
        path: Path,
        event_id: int,
        grouped: bool = True,
        sort_by: str = "place",
        sort_desc: bool = False,
        group_by: str = "heat",
    ) -> Path:
        event = next(e for e in self.repo.list_events() if e.id == event_id)
        swimmers = self.repo.list_swimmers(event_id)
        results_mode = any(self._swimmer_has_result(swimmer) for swimmer in swimmers)
        return self._export_protocol_workbook(
            path=path,
            page_title=self.repo.get_meta("competition_title") or event.name,
            events=[(event.name, swimmers)],
            final_mode=results_mode,
            sort_by=sort_by,
            sort_desc=sort_desc,
            group_by=group_by,
            grouped=grouped,
            show_heat_column=False,
        )

    def export_final_protocol_excel(
        self,
        path: Path,
        grouped: bool = True,
        sort_by: str = "place",
        sort_desc: bool = False,
        group_by: str = "heat",
    ) -> Path:
        events = []
        for event in self.repo.list_events():
            swimmers = self._filter_final_protocol_swimmers(self.repo.list_swimmers(event.id))
            events.append((event.name, swimmers))
        return self._export_protocol_workbook(
            path=path,
            page_title=f"Итоговый протокол {self.repo.get_meta('competition_title') or 'соревнований'}",
            events=events,
            final_mode=True,
            sort_by=sort_by,
            sort_desc=sort_desc,
            group_by=group_by,
            grouped=grouped,
            show_heat_column=False,
        )

    @staticmethod
    def _cell_text_length(value: object) -> int:
        if value is None:
            return 0
        text = str(value).strip()
        if not text:
            return 0
        return max(len(part) for part in text.splitlines())

    def _autosize_protocol_columns(self, ws, headers: list[str], content_lengths: list[int], *, final_mode: bool, show_heat_column: bool) -> None:
        if final_mode:
            limits = [
                (4, 8),   # №
                (18, 36), # Фамилия Имя
                (10, 14), # Год рождения
                (18, 34), # Команда
                (14, 22), # Время
                (6, 10),  # Место
            ]
        elif show_heat_column:
            limits = [
                (6, 10),  # Заплыв
                (6, 10),  # Дорожка
                (18, 36), # Ф. И.
                (10, 14), # Год рождения
                (18, 34), # Команда
                (14, 22), # Заявочное время
            ]
        else:
            limits = [
                (6, 10),  # Дорожка
                (18, 36), # Ф. И.
                (10, 14), # Год рождения
                (18, 34), # Команда
                (14, 22), # Заявочное время
            ]

        measured_widths: list[float] = []
        for index, (header, max_len, (min_width, max_width)) in enumerate(zip(headers, content_lengths, limits), start=1):
            basis = max(max_len, self._cell_text_length(header))
            computed_width = basis * 1.1 + 2
            width = max(min_width, min(max_width, computed_width))
            ws.column_dimensions[get_column_letter(index)].width = width
            measured_widths.append(width)

        # Заполняем доступную ширину листа A4 (fitToWidth=1) и приоритетно расширяем колонку времени.
        target_total_width = 118 if final_mode else (110 if show_heat_column else 104)
        current_total_width = sum(measured_widths)
        if current_total_width >= target_total_width:
            return

        free_space = target_total_width - current_total_width
        time_index = len(headers) - 2 if final_mode or show_heat_column else len(headers) - 1
        time_min, time_max = limits[time_index]
        current_time_width = measured_widths[time_index]
        time_extra = min(free_space * 0.55, max(0, time_max - current_time_width))
        if time_extra > 0:
            measured_widths[time_index] += time_extra
            free_space -= time_extra

        if free_space <= 0:
            for idx, width in enumerate(measured_widths, start=1):
                ws.column_dimensions[get_column_letter(idx)].width = width
            return

        expandable_indices = [
            idx for idx, (_, max_width) in enumerate(limits)
            if idx != time_index and measured_widths[idx] < max_width
        ]
        while free_space > 0.01 and expandable_indices:
            share = free_space / len(expandable_indices)
            next_indices: list[int] = []
            distributed = 0.0
            for idx in expandable_indices:
                max_width = limits[idx][1]
                headroom = max_width - measured_widths[idx]
                delta = min(share, headroom)
                if delta > 0:
                    measured_widths[idx] += delta
                    distributed += delta
                if measured_widths[idx] + 0.01 < max_width:
                    next_indices.append(idx)
            if distributed <= 0:
                break
            free_space -= distributed
            expandable_indices = next_indices

        for idx, width in enumerate(measured_widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width

    def _build_group_title(self, event_name: str, gender_label: str, age_label: str, final_mode: bool) -> str:
        title_parts = [self._base_event_name(event_name)]
        show_gender = not (final_mode and gender_label == "Все")
        if show_gender:
            title_parts.append(gender_label)
        if age_label != "Все возраста" and (gender_label != "Все" or not final_mode):
            title_parts.append(age_label)
        return ", ".join(title_parts).upper()

    def _export_protocol_workbook(
        self,
        path: Path,
        page_title: str,
        events: list[tuple[str, list]],
        final_mode: bool,
        sort_by: str,
        sort_desc: bool,
        group_by: str,
        grouped: bool,
        show_heat_column: bool,
    ) -> Path:
        wb = Workbook()
        ws = wb.active
        ws.title = "Протокол"
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_margins.left = 0.3
        ws.page_margins.right = 0.3
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5
        ws.sheet_view.showGridLines = False

        total_columns = 6 if (final_mode or show_heat_column) else 5

        thin = Side(style="thin", color="CFCFCF")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        title_font = Font(name="Arial", size=16, bold=True)
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        table_header_font = Font(name="Arial", size=11, bold=True)
        body_font = Font(name="Arial", size=10)
        fills = {
            "boys": PatternFill("solid", fgColor="5B9BD5"),
            "girls": PatternFill("solid", fgColor="E06666"),
            "mixed": PatternFill("solid", fgColor="808080"),
        }

        doc_title = page_title if final_mode else (self.repo.get_meta("competition_title") or "Итоговый протокол соревнований")
        row = 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_columns)
        cell = ws.cell(row=row, column=1, value=doc_title)
        cell.font = title_font
        cell.alignment = center
        row += 1

        for meta_value in (self.repo.get_meta("competition_date") or "", self.repo.get_meta("competition_place") or ""):
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_columns)
            cell = ws.cell(row=row, column=1, value=meta_value)
            cell.font = Font(name="Arial", size=11)
            cell.alignment = center
            cell.border = border
            row += 1
        row += 1

        for event_name, swimmers in events:
            groups = self._split_event_into_age_groups(event_name, swimmers)
            for group in groups:
                ranked, places = self._rank_swimmers(group["swimmers"])
                table_title = self._build_group_title(
                    event_name=event_name,
                    gender_label=group["gender_label"],
                    age_label=group["age_label"],
                    final_mode=final_mode,
                )

                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_columns)
                title_cell = ws.cell(row=row, column=1, value=table_title)
                title_cell.font = header_font
                title_cell.alignment = center
                title_cell.fill = fills[group["gender_color"]]
                title_cell.border = border
                ws.row_dimensions[row].height = 24
                row += 1

                headers = (
                    ["№", "Фамилия Имя", "Год рождения", "Команда", "Время", "Место"]
                    if final_mode
                    else (
                        ["Заплыв", "Дорожка", "Ф. И.", "Год рождения", "Команда", "Заявочное время"]
                        if show_heat_column
                        else ["Дорожка", "Ф. И.", "Год рождения", "Команда", "Заявочное время"]
                    )
                )
                content_lengths = [self._cell_text_length(value) for value in headers]
                for column, value in enumerate(headers, start=1):
                    header_cell = ws.cell(row=row, column=column, value=value)
                    header_cell.font = table_header_font
                    header_cell.alignment = center if column != 3 and column != 5 else left
                    header_cell.border = border
                ws.row_dimensions[row].height = 22
                row += 1

                if final_mode:
                    ordered = self._sort_protocol_rows(group["swimmers"], places, sort_by, sort_desc, final_mode=True)
                    for index, swimmer in enumerate(ordered, start=1):
                        values = [
                            index,
                            swimmer.full_name,
                            swimmer.birth_year or "",
                            swimmer.team or "",
                            self._display_swimmer_time(swimmer),
                            places.get(swimmer.id, ""),
                        ]
                        for column, value in enumerate(values, start=1):
                            body_cell = ws.cell(row=row, column=column, value=value)
                            body_cell.font = body_font
                            body_cell.alignment = left if column in {2, 4} else center
                            body_cell.border = border
                            content_lengths[column - 1] = max(content_lengths[column - 1], self._cell_text_length(value))
                        row += 1
                else:
                    if grouped and group_by == "heat" and show_heat_column:
                        ordered = sorted(
                            group["swimmers"],
                            key=lambda s: (s.heat is None, s.heat or 999, s.lane is None, s.lane or 999, s.full_name.lower()),
                        )
                        heat_sizes: dict[int | None, int] = {}
                        for swimmer in ordered:
                            heat_sizes[swimmer.heat] = heat_sizes.get(swimmer.heat, 0) + 1
                        rendered_heats: set[int | None] = set()
                        for swimmer in ordered:
                            start_row = row
                            is_first_heat_row = swimmer.heat not in rendered_heats
                            values = [
                                swimmer.heat or "",
                                swimmer.lane or "",
                                swimmer.full_name,
                                swimmer.birth_year or "",
                                swimmer.team or "",
                                swimmer.seed_time_raw or "",
                            ]
                            for column, value in enumerate(values, start=1):
                                if column == 1 and not is_first_heat_row:
                                    continue
                                body_cell = ws.cell(row=row, column=column, value=value)
                                body_cell.font = body_font
                                body_cell.alignment = left if column in {3, 5} else center
                                body_cell.border = border
                                content_lengths[column - 1] = max(content_lengths[column - 1], self._cell_text_length(value))
                            if is_first_heat_row:
                                rendered_heats.add(swimmer.heat)
                                rowspan = heat_sizes.get(swimmer.heat, 1)
                                if rowspan > 1:
                                    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row + rowspan - 1, end_column=1)
                                ws.cell(row=start_row, column=1).alignment = center
                            row += 1
                    else:
                        ordered = self._sort_protocol_rows(group["swimmers"], places, sort_by, sort_desc, final_mode=False)
                        for swimmer in ordered:
                            values = (
                                [
                                    swimmer.heat or "",
                                    swimmer.lane or "",
                                    swimmer.full_name,
                                    swimmer.birth_year or "",
                                    swimmer.team or "",
                                    swimmer.seed_time_raw or "",
                                ]
                                if show_heat_column
                                else [
                                    swimmer.lane or "",
                                    swimmer.full_name,
                                    swimmer.birth_year or "",
                                    swimmer.team or "",
                                    swimmer.seed_time_raw or "",
                                ]
                            )
                            for column, value in enumerate(values, start=1):
                                body_cell = ws.cell(row=row, column=column, value=value)
                                body_cell.font = body_font
                                body_cell.alignment = (
                                    left if column in ({3, 5} if show_heat_column else {2, 4}) else center
                                )
                                body_cell.border = border
                                content_lengths[column - 1] = max(content_lengths[column - 1], self._cell_text_length(value))
                            row += 1
                self._autosize_protocol_columns(
                    ws,
                    headers,
                    content_lengths,
                    final_mode=final_mode,
                    show_heat_column=show_heat_column,
                )
                row += 1

        output_path = Path(path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        return output_path

    def _build_age_group_table_html(
        self,
        event_name: str,
        swimmers: list,
        age_label: str,
        gender_label: str,
        gender_color: str,
        final_mode: bool,
        sort_by: str,
        sort_desc: bool,
        group_by: str,
        grouped: bool,
        show_heat_column: bool,
    ) -> str:
        ranked, places = self._rank_swimmers(swimmers)
        title = self._build_group_title(
            event_name=event_name,
            gender_label=gender_label,
            age_label=age_label,
            final_mode=final_mode,
        )
        start_protocol_colgroup = (
            "<colgroup><col class='col-2'><col class='col-3'><col class='col-4'><col class='col-5'><col class='col-6'></colgroup>"
            if not show_heat_column and not final_mode
            else "<colgroup><col class='col-1'><col class='col-2'><col class='col-3'><col class='col-4'><col class='col-5'><col class='col-6'></colgroup>"
        )
        total_columns = 5 if not show_heat_column and not final_mode else 6
        parts = [f"<table class='protocol-table'>{start_protocol_colgroup}"]
        parts.append(f"<tr><td class='category-title {gender_color}' colspan='{total_columns}'>{html.escape(title)}</td></tr>")
        if final_mode:
            parts.append("<tr><th class='num'>№</th><th>Фамилия Имя</th><th class='year'>Год рождения</th><th>Команда</th><th class='time'>Время</th><th class='place'>Место</th></tr>")
            ordered = self._sort_protocol_rows(swimmers, places, sort_by, sort_desc, final_mode=True)
            for index, swimmer in enumerate(ordered, start=1):
                parts.append(
                    "<tr>"
                    f"<td class='num'>{index}</td>"
                    f"<td class='name'>{html.escape(swimmer.full_name)}</td>"
                    f"<td class='year'>{swimmer.birth_year or ''}</td>"
                    f"<td>{html.escape(swimmer.team or '')}</td>"
                    f"<td class='time'>{html.escape(self._display_swimmer_time(swimmer))}</td>"
                    f"<td class='place'>{places.get(swimmer.id, '')}</td>"
                    "</tr>"
                )
        else:
            parts.append(
                "<tr><th class='heat'>Заплыв</th><th class='lane'>Дорожка</th><th>Ф. И.</th><th class='year'>Год рождения</th><th>Команда</th><th class='time'>Заявочное время</th></tr>"
                if show_heat_column
                else "<tr><th class='lane'>Дорожка</th><th>Ф. И.</th><th class='year'>Год рождения</th><th>Команда</th><th class='time'>Заявочное время</th></tr>"
            )
            if grouped and group_by == "heat" and show_heat_column:
                ordered = sorted(swimmers, key=lambda s: (s.heat is None, s.heat or 999, s.lane is None, s.lane or 999, s.full_name.lower()))
                heat_sizes: dict[int | None, int] = {}
                for swimmer in ordered:
                    heat_sizes[swimmer.heat] = heat_sizes.get(swimmer.heat, 0) + 1
                rendered_heats: set[int | None] = set()
                for swimmer in ordered:
                    parts.append("<tr>")
                    if swimmer.heat not in rendered_heats:
                        rendered_heats.add(swimmer.heat)
                        rowspan = heat_sizes.get(swimmer.heat, 1)
                        parts.append(f"<td class='heat heat-merged' rowspan='{rowspan}'>{swimmer.heat or ''}</td>")
                    parts.append(
                        f"<td class='lane'>{swimmer.lane or ''}</td>"
                        f"<td class='name'>{html.escape(swimmer.full_name)}</td>"
                        f"<td class='year'>{swimmer.birth_year or ''}</td>"
                        f"<td>{html.escape(swimmer.team or '')}</td>"
                        f"<td class='time'>{html.escape(swimmer.seed_time_raw or '')}</td>"
                        "</tr>"
                    )
            else:
                ordered = self._sort_protocol_rows(swimmers, places, sort_by, sort_desc, final_mode=False)
                for swimmer in ordered:
                    parts.append("<tr>")
                    if show_heat_column:
                        parts.append(f"<td class='heat'>{swimmer.heat or ''}</td>")
                    parts.append(
                        f"<td class='lane'>{swimmer.lane or ''}</td>"
                        f"<td class='name'>{html.escape(swimmer.full_name)}</td>"
                        f"<td class='year'>{swimmer.birth_year or ''}</td>"
                        f"<td>{html.escape(swimmer.team or '')}</td>"
                        f"<td class='time'>{html.escape(swimmer.seed_time_raw or '')}</td>"
                        "</tr>"
                    )
        parts.append("</table>")
        return "".join(parts)

    def _rank_swimmers(self, swimmers: list) -> tuple[list, dict[int, int]]:
        active = [s for s in swimmers if s.status != "DNS"]
        ranked = sorted(active, key=lambda s: (self._ranking_time_cs(s) is None, self._ranking_time_cs(s) or 99999999, s.full_name.lower()))
        places: dict[int, int] = {}
        last_time_cs = None
        for idx, swimmer in enumerate(ranked, start=1):
            current_time_cs = self._ranking_time_cs(swimmer)
            if current_time_cs is None:
                continue
            if current_time_cs != last_time_cs:
                places[swimmer.id] = idx
                last_time_cs = current_time_cs
                continue
            places[swimmer.id] = places[ranked[idx - 2].id]
        return ranked, places

    def _ranking_time_cs(self, swimmer) -> int | None:
        return swimmer.result_time_cs

    def _display_swimmer_time(self, swimmer) -> str:
        return swimmer.result_time_raw or swimmer.result_mark or ""

    def _swimmer_has_result(self, swimmer) -> bool:
        return swimmer.result_time_cs is not None or bool((swimmer.result_mark or "").strip())

    def _sort_protocol_rows(self, swimmers: list, places: dict[int, int], sort_by: str, sort_desc: bool, final_mode: bool) -> list:
        def sort_key(s):
            if sort_by == "id":
                return (s.id,)
            if sort_by == "team":
                team = (s.team or "").strip()
                return (team == "", team.lower(), s.full_name.lower())
            if sort_by == "birth_year":
                return (s.birth_year is None, s.birth_year or 0, s.full_name.lower())
            if sort_by == "seed_time":
                return (s.seed_time_cs is None, s.seed_time_cs or 99999999, s.full_name.lower())
            if sort_by == "result_time":
                return (s.result_time_cs is None, s.result_time_cs or 99999999, s.full_name.lower())
            if sort_by == "heat":
                return (s.heat is None, s.heat or 999, s.lane is None, s.lane or 999, s.full_name.lower())
            if sort_by == "lane":
                return (s.lane is None, s.lane or 999, s.heat is None, s.heat or 999, s.full_name.lower())
            if sort_by == "status":
                return ((s.status or "").lower(), s.full_name.lower())
            if sort_by == "mark":
                mark = (s.result_mark or "").strip()
                return (mark == "", mark, s.full_name.lower())
            if sort_by == "full_name":
                return (s.full_name.lower(),)
            if final_mode:
                return (places.get(s.id, 10**9), s.full_name.lower())
            return (s.heat is None, s.heat or 999, s.lane is None, s.lane or 999, s.full_name.lower())

        return sorted(swimmers, key=sort_key, reverse=sort_desc)

    def _filter_final_protocol_swimmers(self, swimmers: list) -> list:
        disallowed_marks = {"DNS", "DQ"}
        return [
            s
            for s in swimmers
            if s.status != "DNS" and ((s.result_mark or "").strip().upper() not in disallowed_marks)
        ]
