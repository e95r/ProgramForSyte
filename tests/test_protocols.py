from pathlib import Path

from openpyxl import load_workbook

from core.service import MeetService


def test_event_protocol_uses_start_heats_and_lanes(tmp_path: Path):
    service = MeetService(tmp_path)
    try:
        service.repo.set_meta("competition_title", 'Открытый турнир по плаванию «АКВАДОН»')
        service.repo.set_meta("competition_date", "25.04.2026")
        service.repo.set_meta("competition_place", "Донской, Тульская область")
        event_id = service.repo.upsert_event("100 брасс, мужчины все")
        service.repo.add_swimmers(
            event_id,
            [
                {"full_name": "B", "heat": 1, "lane": 1, "birth_year": 2013, "team": "Sharks", "seed_time_raw": "00:35:00", "seed_time_cs": 3500},
                {"full_name": "A", "heat": 1, "lane": 2, "birth_year": 2012, "team": "Dolphins", "seed_time_raw": "00:33:00", "seed_time_cs": 3300},
            ],
        )

        html = service.build_event_protocol(event_id, grouped=True)
        assert "<div class='doc-title'>Открытый турнир по плаванию «АКВАДОН»</div>" in html
        assert "25.04.2026" in html
        assert "Донской, Тульская область" in html
        assert "100 БРАСС, МУЖЧИНЫ" in html
        assert "100 БРАСС, МУЖЧИНЫ, ВСЕ ВОЗРАСТА" not in html
        assert ">Заплыв<" not in html
        assert "rowspan='2'>1</td>" not in html
        assert "<td class='lane'>1</td>" in html
        assert "<td class='lane'>2</td>" in html
        assert "Соревнование:" not in html
    finally:
        service.close()


def test_import_startlist_keeps_combined_heat_and_lane_from_excel(tmp_path: Path):
    from openpyxl import Workbook

    source = tmp_path / "source.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "50m"
    ws.append(["ФИО", "Заявочное время", "Заплыв/дорожка"])
    ws.append(["Slow", "00:40:00", "6/3"])
    ws.append(["Fast", "00:30:00", "5/1"])
    wb.save(source)

    service = MeetService(tmp_path)
    try:
        service.import_startlist(source)
        event_id = service.repo.list_events()[0].id
        swimmers = service.repo.list_swimmers(event_id)
        assert [(s.full_name, s.heat, s.lane) for s in swimmers] == [("Fast", 5, 1), ("Slow", 6, 3)]
    finally:
        service.close()


def test_import_startlist_removes_trailing_all_from_gender_event_titles(tmp_path: Path):
    from openpyxl import Workbook

    source = tmp_path / "gendered-source.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Старт"
    ws.append(["Турнир"])
    ws.append(["100 брасс, женщины все"])
    ws.append(["Заплыв", "Дорожка", "Ф. И.", "Год рождения", "Команда", "Заявочное время"])
    ws.append([1, 3, "Белова Екатерина", 2010, "Team", "00:40:00"])
    wb.save(source)

    service = MeetService(tmp_path)
    try:
        service.import_startlist(source)
        event = service.repo.list_events()[0]
        assert event.name == "100 брасс, женщины"
    finally:
        service.close()


def test_import_startlist_keeps_grouped_protocol_heats_and_lanes(tmp_path: Path):
    from openpyxl import Workbook

    source = tmp_path / "grouped-source.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Старт"
    ws.append(["Турнир"])
    ws.append(["100 — Брасс все"])
    ws.append(["Заплыв", "Дорожка", "Ф. И.", "Год рождения", "Команда", "Заявочное время"])
    ws.append([2, 3, "Slow", 2010, "Team", "00:40:00"])
    ws.append([None, 4, "Fast", 2011, "Team", "00:30:00"])
    wb.save(source)

    service = MeetService(tmp_path)
    try:
        service.import_startlist(source)
        event = service.repo.list_events()[0]
        swimmers = service.repo.list_swimmers(event.id)
        assert event.name == "100 — Брасс"
        assert [(s.full_name, s.heat, s.lane) for s in swimmers] == [("Slow", 2, 3), ("Fast", 2, 4)]
    finally:
        service.close()


def test_import_startlist_keeps_combined_protocol_and_infers_lane_count(tmp_path: Path):
    from openpyxl import Workbook

    source = tmp_path / "combined-six-lanes.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "100 — Брасс все"
    ws.append(["ФИО", "Заявочное время", "Заплыв/дорожка"])
    ws.append(["Жуков Илья", "20.31.20", "1/1"])
    ws.append(["Кузнецов Юрий", "5.16.05", "1/2"])
    ws.append(["Андреев Александр", "1.38.53", "1/3"])
    ws.append(["Белова Екатерина", "4.15.32", "1/4"])
    ws.append(["Козлов Андрей", "6.12.06", "1/5"])
    ws.append(["Андреева Алина", "1.35.11", "2/1"])
    ws.append(["Кузнецова Ольга", "1.06.42", "2/2"])
    ws.append(["Васильев Григорий", "1.01.35", "2/3"])
    ws.append(["Андреева Дарья", "1.01.53", "2/4"])
    ws.append(["Гаврилова Вероника", "1.34.26", "2/5"])
    ws.append(["Жуков Максим", "1.36.02", "2/6"])
    wb.save(source)

    service = MeetService(tmp_path)
    try:
        service.import_startlist(source)
        event = service.repo.list_events()[0]
        swimmers = service.repo.list_swimmers(event.id)

        assert event.lanes_count == 6
        assert [(s.full_name, s.heat, s.lane) for s in swimmers] == [
            ("Жуков Илья", 1, 1),
            ("Кузнецов Юрий", 1, 2),
            ("Андреев Александр", 1, 3),
            ("Белова Екатерина", 1, 4),
            ("Козлов Андрей", 1, 5),
            ("Андреева Алина", 2, 1),
            ("Кузнецова Ольга", 2, 2),
            ("Васильев Григорий", 2, 3),
            ("Андреева Дарья", 2, 4),
            ("Гаврилова Вероника", 2, 5),
            ("Жуков Максим", 2, 6),
        ]
    finally:
        service.close()


def test_final_protocol_contains_all_events_and_metadata(tmp_path: Path):
    service = MeetService(tmp_path)
    try:
        service.repo.set_meta("competition_title", 'открытого турнира по плаванию "АКВАДОН"')
        service.repo.set_meta("competition_date", "20 декабря 2025")
        service.repo.set_meta("competition_place", "Тульская обл. г.Донской МБУ ДСК")
        e1 = service.repo.upsert_event("50m")
        e2 = service.repo.upsert_event("100m")
        service.repo.add_swimmers(e1, [{"full_name": "A", "heat": 1, "lane": 1}])
        service.repo.add_swimmers(e2, [{"full_name": "B", "heat": 1, "lane": 2}])

        html = service.build_final_protocol(grouped=True)
        assert "Итоговый протокол открытого турнира по плаванию &quot;АКВАДОН&quot;" in html
        assert "20 декабря 2025" in html
        assert "Тульская обл. г.Донской МБУ ДСК" in html
        assert "50M, ВСЕ, ВСЕ ВОЗРАСТА" in html
        assert "100M, ВСЕ, ВСЕ ВОЗРАСТА" in html
        assert "size: A4" in html
    finally:
        service.close()


def test_final_protocol_excludes_dns_and_dq(tmp_path: Path):
    service = MeetService(tmp_path)
    try:
        event_id = service.repo.upsert_event("200m")
        service.repo.add_swimmers(
            event_id,
            [
                {"full_name": "A", "status": "DNS", "result_mark": ""},
                {"full_name": "B", "status": "OK", "result_mark": "DQ"},
                {"full_name": "C", "status": "OK", "result_mark": ""},
            ],
        )

        html = service.build_final_protocol(grouped=False)
        assert ">C</td>" in html
        assert ">A</td>" not in html
        assert ">B</td>" not in html
    finally:
        service.close()


def test_final_protocol_uses_only_result_time_not_seed_time(tmp_path: Path):
    service = MeetService(tmp_path)
    try:
        event_id = service.repo.upsert_event("200m")
        service.repo.add_swimmers(
            event_id,
            [
                {"full_name": "WithResult", "seed_time_raw": "00:40:00", "seed_time_cs": 4000, "result_time_raw": "00:38:50", "result_time_cs": 3850},
                {"full_name": "NoResult", "seed_time_raw": "00:41:00", "seed_time_cs": 4100},
            ],
        )

        html = service.build_final_protocol(grouped=False)
        assert ">00:38:50<" in html
        assert ">00:40:00<" not in html
        assert ">00:41:00<" not in html
    finally:
        service.close()


def test_final_protocol_splits_by_age_groups_and_colors_by_gender(tmp_path: Path):
    service = MeetService(tmp_path)
    try:
        service.repo.set_meta(
            "age_groups",
            '[{"index": 1, "label": "2010 и старше", "min_year": null, "max_year": 2010}, '
            '{"index": 2, "label": "2011-2012", "min_year": 2011, "max_year": 2012}, '
            '{"index": 3, "label": "2013-2014", "min_year": 2013, "max_year": 2014}]',
        )
        event_id = service.repo.upsert_event("100 комплексное плавание, мужчины все")
        service.repo.add_swimmers(
            event_id,
            [
                {"full_name": "Old", "birth_year": 2010, "team": "A", "result_time_raw": "01:02:29", "result_time_cs": 6229},
                {"full_name": "Mid", "birth_year": 2011, "team": "B", "result_time_raw": "01:03:40", "result_time_cs": 6340},
                {"full_name": "Young", "birth_year": 2014, "team": "C", "result_time_raw": "01:07:50", "result_time_cs": 6750},
            ],
        )

        html = service.build_final_protocol()
        assert "100 КОМПЛЕКСНОЕ ПЛАВАНИЕ, МУЖЧИНЫ, 2010 И СТАРШЕ" in html
        assert "100 КОМПЛЕКСНОЕ ПЛАВАНИЕ, МУЖЧИНЫ, 2011-2012" in html
        assert "100 КОМПЛЕКСНОЕ ПЛАВАНИЕ, МУЖЧИНЫ, 2013-2014" in html
        assert "category-title boys" in html
        assert html.count("<table class='protocol-table'>") == 3
    finally:
        service.close()


def test_event_protocol_uses_women_gender_title_without_all_ages_suffix(tmp_path: Path):
    service = MeetService(tmp_path)
    try:
        event_id = service.repo.upsert_event("100 брасс, женщины все")
        service.repo.add_swimmers(
            event_id,
            [
                {"full_name": "A", "heat": 1, "lane": 1, "birth_year": 2012, "team": "Team", "seed_time_raw": "00:33:00", "seed_time_cs": 3300},
            ],
        )

        html = service.build_event_protocol(event_id, grouped=True)

        assert "100 БРАСС, ЖЕНЩИНЫ" in html
        assert "100 БРАСС, ЖЕНЩИНЫ, ВСЕ ВОЗРАСТА" not in html
    finally:
        service.close()


def test_event_protocol_with_results_uses_competition_header_and_places(tmp_path: Path):
    service = MeetService(tmp_path)
    try:
        service.repo.set_meta("competition_title", "Открытый турнир по плаванию Аквадон")
        service.repo.set_meta("competition_date", "25.04.2026")
        service.repo.set_meta("competition_place", "Донской, Тульская область")
        service.repo.set_meta(
            "age_groups",
            '[{"index": 1, "label": "2010 и старше", "min_year": null, "max_year": 2010}]',
        )
        event_id = service.repo.upsert_event("100 брасс, женщины")
        service.repo.add_swimmers(
            event_id,
            [
                {"full_name": "Белова Екатерина", "birth_year": 2010, "team": "Команда 1", "result_time_raw": "1.38.47", "result_time_cs": 9847},
                {"full_name": "Андреева Алина", "birth_year": 2000, "team": "Команда 2", "result_time_raw": "1.31.43", "result_time_cs": 9143},
                {"full_name": "Гаврилова Вероника", "birth_year": 2005, "team": "Команда 3", "result_time_raw": "1.34.53", "result_time_cs": 9453},
            ],
        )

        html = service.build_event_protocol(event_id)

        assert "<div class='doc-title'>Открытый турнир по плаванию Аквадон</div>" in html
        assert "100 БРАСС, ЖЕНЩИНЫ, 2010 И СТАРШЕ" in html
        assert "<th class='time'>Время</th><th class='place'>Место</th>" in html
        assert html.index("Андреева Алина") < html.index("Гаврилова Вероника") < html.index("Белова Екатерина")
        assert html.count("<td class='place'>1</td>") == 1
        assert html.count("<td class='place'>2</td>") == 1
        assert html.count("<td class='place'>3</td>") == 1
    finally:
        service.close()


def test_final_protocol_place_sort_desc(tmp_path: Path):
    service = MeetService(tmp_path)
    try:
        event_id = service.repo.upsert_event("50m")
        service.repo.add_swimmers(
            event_id,
            [
                {"full_name": "Fast", "result_time_raw": "00:30:00", "result_time_cs": 3000},
                {"full_name": "Slow", "result_time_raw": "00:35:00", "result_time_cs": 3500},
            ],
        )

        html = service.build_final_protocol(grouped=False, sort_by="place", sort_desc=True)
        assert html.index("Slow") < html.index("Fast")
    finally:
        service.close()


def test_final_protocol_tie_places_skip_next_place(tmp_path: Path):
    service = MeetService(tmp_path)
    try:
        event_id = service.repo.upsert_event("50m")
        service.repo.add_swimmers(
            event_id,
            [
                {"full_name": "A", "result_time_raw": "00:30:00", "result_time_cs": 3000},
                {"full_name": "B", "result_time_raw": "00:31:00", "result_time_cs": 3100},
                {"full_name": "C", "result_time_raw": "00:31:00", "result_time_cs": 3100},
                {"full_name": "D", "result_time_raw": "00:32:00", "result_time_cs": 3200},
            ],
        )

        html = service.build_final_protocol(grouped=False, sort_by="place")
        assert "<td class='place'>1</td>" in html
        assert html.count("<td class='place'>2</td>") == 2
        assert "<td class='place'>4</td>" in html
    finally:
        service.close()


def test_final_protocol_wraps_content_to_a4_width_and_uses_fixed_table_columns(tmp_path: Path):
    service = MeetService(tmp_path)
    try:
        event_id = service.repo.upsert_event("50m")
        service.repo.add_swimmers(event_id, [{"full_name": "A", "result_time_raw": "00:30:00", "result_time_cs": 3000}])

        html = service.build_final_protocol(grouped=False)

        assert ".page { width: 186mm; max-width: 186mm; margin: 0 auto; box-sizing: border-box; }" in html
        assert "<table class='protocol-table'><colgroup><col class='col-1'><col class='col-2'><col class='col-3'><col class='col-4'><col class='col-5'><col class='col-6'></colgroup>" in html
    finally:
        service.close()




def test_event_protocol_grouped_by_heat_can_be_exported_to_excel(tmp_path: Path):
    service = MeetService(tmp_path)
    try:
        service.repo.set_meta("competition_title", 'Открытый турнир по плаванию «АКВАДОН»')
        service.repo.set_meta("competition_date", "25.04.2026")
        service.repo.set_meta("competition_place", "Донской, Тульская область")
        event_id = service.repo.upsert_event("100 брасс, мужчины все")
        service.repo.add_swimmers(
            event_id,
            [
                {"full_name": "Андреев Андрей", "heat": 1, "lane": 3, "birth_year": 2008, "team": "Команда 1", "seed_time_raw": "01:01:01", "seed_time_cs": 6101},
                {"full_name": "Иванов Иван", "heat": 1, "lane": 4, "birth_year": 2010, "team": "Команда 2", "seed_time_raw": "01:02:02", "seed_time_cs": 6202},
            ],
        )

        target = tmp_path / "event-protocol.xlsx"
        saved = service.export_event_protocol_excel(target, event_id, grouped=True, group_by="heat")

        assert saved == target
        assert target.exists()
        wb = load_workbook(target)
        ws = wb.active
        assert "A7:A8" not in {str(range_ref) for range_ref in ws.merged_cells.ranges}
        assert ws["A6"].value == "Дорожка"
        assert ws["A7"].value == 3
        assert ws["A8"].value == 4
        assert ws["B8"].value == "Иванов Иван"
    finally:
        service.close()


def test_final_protocol_can_be_exported_to_excel_with_a4_setup(tmp_path: Path):
    service = MeetService(tmp_path)
    try:
        service.repo.set_meta("competition_title", 'Открытый турнир по плаванию «АКВАДОН»')
        service.repo.set_meta("competition_date", "25.04.2026")
        service.repo.set_meta("competition_place", "Донской, Тульская область")
        event_id = service.repo.upsert_event("100 брасс, мужчины все")
        service.repo.add_swimmers(
            event_id,
            [
                {"full_name": "Андреев Андрей", "birth_year": 2008, "team": "Команда 1", "result_time_raw": "01:01:01", "result_time_cs": 6101},
                {"full_name": "Иванов Иван", "birth_year": 2010, "team": "Команда 2", "result_time_raw": "01:02:02", "result_time_cs": 6202},
            ],
        )

        target = tmp_path / "final-protocol.xlsx"
        saved = service.export_final_protocol_excel(target, grouped=False)

        assert saved == target
        assert target.exists()
        wb = load_workbook(target)
        ws = wb.active
        assert str(ws.page_setup.paperSize) == ws.PAPERSIZE_A4
        assert ws.page_setup.fitToWidth == 1
        assert ws["A1"].value == "Итоговый протокол Открытый турнир по плаванию «АКВАДОН»"
        assert ws["A5"].value == "100 БРАСС, МУЖЧИНЫ"
        assert ws["B7"].value == "Андреев Андрей"
        assert ws["F8"].value == 2
        assert ws.column_dimensions["B"].width > ws.column_dimensions["C"].width
        assert ws.column_dimensions["D"].width > ws.column_dimensions["C"].width
        assert ws.column_dimensions["A"].width < ws.column_dimensions["B"].width
        assert ws.column_dimensions["F"].width < ws.column_dimensions["B"].width
    finally:
        service.close()
