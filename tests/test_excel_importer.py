from pathlib import Path

import pytest
from openpyxl import Workbook

from core.excel_importer import ExcelImportError, extract_meet_metadata, import_excel


def test_import_excel_rejects_legacy_xls(tmp_path: Path):
    legacy_file = tmp_path / "startlist.xls"
    legacy_file.write_text("legacy excel")

    with pytest.raises(ExcelImportError, match=".xls"):
        import_excel(legacy_file)


def test_import_excel_parses_birth_year_as_text(tmp_path: Path):
    file_path = tmp_path / "startlist.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "50m"
    ws.append(["ФИО", "Год рождения", "Заплыв/дорожка"])
    ws.append(["Иванов Иван", "2012 г.", "6/3"])
    wb.save(file_path)

    data = import_excel(file_path)
    swimmer = data["50m"][0]
    assert swimmer["birth_year"] == 2012


def test_import_excel_parses_grouped_heats_from_protocol_blocks(tmp_path: Path):
    file_path = tmp_path / "grouped-startlist.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Лист 1"
    ws.append(['Турнир по плаванию "АКВАДОН"'])
    ws.append([])
    ws.append(["100 — Брасс все"])
    ws.append(["Заплыв", "Дорожка", "Ф. И.", "Год рождения", "Команда", "Заявочное время"])
    ws.append([1, 1, "Жуков Илья", 2007, "Команда для теста", "20.31.20"])
    ws.append([None, 2, "Кузнецов Юрий", 2009, "Команда для теста", "5.16.05"])
    ws.append([2, 1, "Андреева Алма", 2000, "Команда для теста", "1.35.11"])
    ws.append([None, 2, "Кузнецов Ольга", 2010, "Команда для теста", "1.06.42"])
    ws.append([])
    ws.append(["100 — Вольный стиль девочки 2010-2012"])
    ws.append(["Заплыв", "Дорожка", "Ф. И.", "Год рождения", "Команда", "Заявочное время"])
    ws.append([1, 2, "Васильев Григорий", 2012, "Команда для теста", "10.41.25"])
    ws.append([None, 4, "Жуков Илья", 2007, "Команда для теста", "1.35.42"])
    wb.save(file_path)

    data = import_excel(file_path)

    assert list(data) == ["100 — Брасс все", "100 — Вольный стиль девочки 2010-2012"]
    assert [(s["full_name"], s["heat"], s["lane"]) for s in data["100 — Брасс все"]] == [
        ("Жуков Илья", 1, 1),
        ("Кузнецов Юрий", 1, 2),
        ("Андреева Алма", 2, 1),
        ("Кузнецов Ольга", 2, 2),
    ]
    assert data["100 — Вольный стиль девочки 2010-2012"][0]["birth_year"] == 2012


def test_import_excel_uses_relay_title_instead_of_protocol_sheet_name(tmp_path: Path):
    file_path = tmp_path / "relay-startlist.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Протокол"
    ws.append(["4X50 — Вольный стиль все"])
    ws.append(["Заплыв", "Дорожка", "Ф. И.", "Год рождения", "Команда", "Заявочное время"])
    ws.append([1, 1, "Белова Екатерина", 2010, "Команда для теста", "10.13.84"])
    ws.append([None, 2, "Кузнецова Ольга", 2010, "Команда для теста", "3.54.10"])
    ws.append([None, 3, "Жуков Илья", 2007, "Команда для теста", "1.53.15"])
    ws.append([None, 4, "Кузнецов Юрий", 2009, "Команда для теста", "3.45.43"])
    ws.append([2, 1, "Андреева Алина", 2000, "Команда для теста", "1.32.85"])
    ws.append([None, 2, "Жуков Максим", 2006, "Команда для теста", "0.54.34"])
    wb.save(file_path)

    data = import_excel(file_path)

    assert list(data) == ["4X50 — Вольный стиль все"]
    assert [(s["full_name"], s["heat"], s["lane"]) for s in data["4X50 — Вольный стиль все"]] == [
        ("Белова Екатерина", 1, 1),
        ("Кузнецова Ольга", 1, 2),
        ("Жуков Илья", 1, 3),
        ("Кузнецов Юрий", 1, 4),
        ("Андреева Алина", 2, 1),
        ("Жуков Максим", 2, 2),
    ]


def test_extract_meet_metadata_reads_competition_details_and_age_groups(tmp_path: Path):
    file_path = tmp_path / "meta-startlist.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Старт"
    ws.append(["Открытый турнир по плаванию «АКВАДОН»"])
    ws.append(["Соревнование: Открытый турнир по плаванию «АКВАДОН»"])
    ws.append(["Дата проведения: 25.04.2026"])
    ws.append(["Место проведения: Донской, Тульская область, г. Донской, мкр. Центральный"])
    ws.append([])
    ws.append(["Возрастные группы:"])
    ws.append(["Группа 1 — 2010 и старше"])
    ws.append(["Группа 2 — 2011-2012"])
    ws.append(["Группа 3 — 2013-2014"])
    ws.append(["Группа 4 — 2015 и младше"])
    ws.append([])
    ws.append(["Эстафетное плавание:"])
    ws.append(["Группа 1 — 2010 и старше"])
    wb.save(file_path)

    meta = extract_meet_metadata(file_path)

    assert meta["competition_title"] == "Открытый турнир по плаванию «АКВАДОН»"
    assert meta["competition_date"] == "25.04.2026"
    assert meta["competition_place"].startswith("Донской")
    assert meta["age_groups"] == [
        {"index": 1, "label": "2010 и старше", "min_year": None, "max_year": 2010},
        {"index": 2, "label": "2011-2012", "min_year": 2011, "max_year": 2012},
        {"index": 3, "label": "2013-2014", "min_year": 2013, "max_year": 2014},
        {"index": 4, "label": "2015 и младше", "min_year": 2015, "max_year": None},
    ]
    assert meta["relay_age_groups"] == [
        {"index": 1, "label": "2010 и старше", "min_year": None, "max_year": 2010}
    ]
